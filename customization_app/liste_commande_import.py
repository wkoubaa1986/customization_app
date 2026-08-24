"""
Backend de l'outil « Liste Commande Import » (demande de cotation fournisseur).

Créée depuis la sélection de la page Prévision Import, la liste est un DocType
persistant (Liste Commande Import + table enfant Article). Ce module fournit :
  - create_from_selection : créer/compléter une liste depuis la sélection ;
  - ai_improve_descriptions / ai_translate : amélioration et traduction des
    descriptions via OpenAI (clé lue dans Raven Settings) ;
  - download_pdf / download_excel : demande de cotation à envoyer au fournisseur.

Réservé à un seul utilisateur (même garde que Prévision Import).
"""

import base64
import html as html_lib
import json
import mimetypes
import os
import re

import frappe
from frappe import _
from frappe.utils import cint, flt, nowdate

ALLOWED_USER = "koubaawassim@gmail.com"

LANG_NAMES = {
    "Français": "French",
    "English": "English",
    "Deutsch": "German",
    "العربية": "Arabic",
    "Español": "Spanish",
}
RTL_LANGS = {"العربية"}

DOCTYPE = "Liste Commande Import"


def _guard():
    if frappe.session.user != ALLOWED_USER:
        frappe.throw(_("Accès réservé."), frappe.PermissionError)


def _strip_html(html):
    if not html:
        return ""
    text = re.sub(r"<br\s*/?>", "\n", html)
    text = re.sub(r"</p>", "\n", text)
    text = re.sub(r"<[^>]+>", "", text)
    return html_lib.unescape(text).strip()


# ---------------------------------------------------------------- création

@frappe.whitelist()
def create_from_selection(items, target=None, titre=None):
    """items = JSON [{item_code, qty}] ; target = liste Brouillon existante à
    compléter, sinon nouvelle liste avec `titre`. Merge : si l'article existe
    déjà dans la liste, la qty transmise remplace l'ancienne."""
    _guard()
    rows = json.loads(items) if isinstance(items, str) else items
    if not rows:
        frappe.throw(_("Aucun article sélectionné."))

    if target:
        doc = frappe.get_doc(DOCTYPE, target)
        if doc.statut != "Brouillon":
            frappe.throw(_("La liste {0} n'est plus en Brouillon.").format(target))
    else:
        doc = frappe.new_doc(DOCTYPE)
        doc.titre = (titre or "").strip() or f"Commande import {nowdate()}"

    existing = {r.item_code: r for r in doc.articles}
    for it in rows:
        code = it.get("item_code")
        qty = flt(it.get("qty"))
        if not code:
            continue
        if code in existing:
            existing[code].qty = qty
            continue
        item = frappe.db.get_value(
            "Item", code,
            ["item_name", "description", "image", "stock_uom", "custom_volume_m3",
             "item_group"],
            as_dict=True) or frappe._dict()
        doc.append("articles", {
            "item_code": code,
            "item_name": item.get("item_name") or code,
            "item_group": item.get("item_group") or "",
            "qty": qty,
            "uom": item.get("stock_uom") or "",
            "description": item.get("description") or "",
            "image": item.get("image") or "",
            "volume_unitaire_m3": flt(item.get("custom_volume_m3")),
        })

    doc.save(ignore_permissions=True)
    frappe.db.commit()
    return {"name": doc.name, "nb_articles": doc.nb_articles}


# ---------------------------------------------------------------- IA (OpenAI)

def _ai_setting(fieldname):
    """Lit un champ du Single « AI Settings » (créé par woocommerce_fusion)."""
    for dt in ("AI Settings", "AI settings"):
        if frappe.db.exists("DocType", dt):
            try:
                val = getattr(frappe.get_cached_doc(dt), fieldname, None)
                if val is not None and str(val).strip():
                    return str(val).strip()
            except Exception:
                pass
            return None
    return None


def _openai_client():
    api_key = (_ai_setting("openai_api_key")
               or frappe.conf.get("openai_api_key")
               or frappe.conf.get("lci_openai_api_key"))
    if not api_key:
        frappe.throw(_("Aucune clé OpenAI configurée (AI Settings > openai_api_key, "
                       "ou site_config openai_api_key)."))
    from openai import OpenAI
    return OpenAI(api_key=api_key)


def _model():
    return (frappe.conf.get("lci_openai_model")
            or _ai_setting("open_ai_model") or _ai_setting("openai_model")
            or "gpt-4o-mini")


def _temperature():
    try:
        return float(_ai_setting("open_ai_temperature") or 0.2)
    except Exception:
        return 0.2


def _chat_json(system, user):
    """Appel completion → JSON object. Lève une erreur claire en cas d'échec."""
    client = _openai_client()
    params = {
        "model": _model(),
        "messages": [{"role": "system", "content": system},
                     {"role": "user", "content": user}],
        "response_format": {"type": "json_object"},
    }
    try:
        resp = client.chat.completions.create(temperature=_temperature(), **params)
    except Exception as e:
        # certains modèles (raisonneurs) refusent une température custom
        if "temperature" in str(e):
            resp = client.chat.completions.create(**params)
        else:
            raise
    content = resp.choices[0].message.content
    try:
        return json.loads(content)
    except Exception:
        frappe.throw(_("Réponse IA illisible : {0}").format(content[:300]))


def _target_rows(doc, row_names):
    names = None
    if row_names:
        names = set(json.loads(row_names) if isinstance(row_names, str) else row_names)
    return [r for r in doc.articles if (names is None or r.name in names)]


AI_CHUNK = 8  # lignes par appel IA — les grosses listes sont traitées par lots


def _chunks(lst, size):
    for i in range(0, len(lst), size):
        yield lst[i:i + size]


@frappe.whitelist()
def ai_improve_descriptions(docname, row_names=None):
    """Réécrit la description technique de chaque ligne ciblée (batch, 1 appel)."""
    _guard()
    doc = frappe.get_doc(DOCTYPE, docname)
    rows = [r for r in _target_rows(doc, row_names)
            if _strip_html(r.description) or r.item_name]
    if not rows:
        frappe.throw(_("Aucune ligne avec description à améliorer."))

    system = ("Tu es un rédacteur technique pour une société de traitement d'eau "
              "(osmoseurs, filtres, adoucisseurs, pompes). Pour chaque article : "
              "1) réécris une DÉSIGNATION courte et claire (max 60 caractères, "
              "nom commercial précis) ; 2) réécris une DESCRIPTION produit sous "
              "forme de LISTE À PUCES, en français, adaptée à une demande de "
              "cotation fournisseur. RÈGLE STRICTE : UNE SEULE caractéristique "
              "par puce — jamais deux caractéristiques réunies par « et », une "
              "virgule ou un point-virgule ; si la source en regroupe "
              "plusieurs, ÉCLATE-les en puces distinctes (autant de puces que "
              "nécessaire). Exemples : « Réservoir 3,2 G, purge manuelle et "
              "socle métallique » devient 3 puces : « • Réservoir : 3,2 G », "
              "« • Purge manuelle », « • Socle métallique ». Chaque ligne "
              "commence par « • » et les puces sont séparées par \\n (aucune "
              "phrase hors des puces). N'invente aucune caractéristique "
              "absente. Réponds en "
              'JSON: {"0": {"designation": "...", "description": "..."}, "1": {...}}')

    updated = []
    for chunk in _chunks(rows, AI_CHUNK):
        payload = {str(i): {"designation": r.item_name or r.item_code or "",
                            "description": _strip_html(r.description) or "(aucune description)"}
                   for i, r in enumerate(chunk)}
        result = _chat_json(system, json.dumps(payload, ensure_ascii=False))
        for i, r in enumerate(chunk):
            entry = result.get(str(i)) or {}
            if isinstance(entry, str):
                entry = {"description": entry}
            new_name = (entry.get("designation") or "").strip()
            new_desc = (entry.get("description") or "").strip()
            if new_name:
                r.item_name = new_name
            if new_desc:
                # puces « • » une par ligne → HTML <br> (Text Editor) ;
                # _strip_html les reconvertit en \n pour le PDF/Excel
                lines = [l.strip() for l in new_desc.splitlines() if l.strip()]
                r.description = "<br>".join(html_lib.escape(l) for l in lines)
            if new_name or new_desc:
                updated.append(r.item_code or r.item_name)
    doc.save(ignore_permissions=True)
    frappe.db.commit()
    return {"updated": updated}


@frappe.whitelist()
def ai_translate(docname, row_names=None):
    """Traduit la description de chaque ligne vers la langue cible du document."""
    _guard()
    doc = frappe.get_doc(DOCTYPE, docname)
    lang = LANG_NAMES.get(doc.langue_cible or "English", "English")
    rows = [r for r in _target_rows(doc, row_names)
            if _strip_html(r.description) or (r.item_name or "").strip()]
    if not rows:
        frappe.throw(_("Aucune ligne avec désignation ou description à traduire."))

    system = (f"Translate each product designation and description to {lang}. "
              "Keep technical terms, units and product codes accurate. "
              "Preserve the formatting exactly: keep line breaks and bullet "
              "points (lines starting with •) as in the source. "
              "Professional tone for a supplier quotation request. Respond in "
              'JSON: {"0": {"designation": "...", "description": "..."}, "1": {...}}')

    updated = []
    for chunk in _chunks(rows, AI_CHUNK):
        payload = {str(i): {"designation": r.item_name or r.item_code or "",
                            "description": _strip_html(r.description)}
                   for i, r in enumerate(chunk)}
        result = _chat_json(system, json.dumps(payload, ensure_ascii=False))
        for i, r in enumerate(chunk):
            entry = result.get(str(i)) or {}
            if isinstance(entry, str):
                entry = {"description": entry}
            tr_name = (entry.get("designation") or "").strip()
            tr_desc = (entry.get("description") or "").strip()
            if tr_name:
                r.item_name_traduit = tr_name
            if tr_desc:
                r.description_traduite = tr_desc
            if tr_name or tr_desc:
                updated.append(r.item_code or r.item_name)

    # traduire aussi les noms des articles ADDITIONNELS embarqués sur les lignes
    rows_with_adds = []
    for r in _target_rows(doc, row_names):
        adds = _row_adds(r)
        if adds:
            rows_with_adds.append((r, adds))
    flat = [(r, adds, j) for r, adds in rows_with_adds
            for j in range(len(adds)) if (adds[j].get("item_name") or "").strip()]
    if flat:
        system_adds = (f"Translate each product name to {lang}. Keep technical "
                       "terms, units and product codes accurate. Respond in "
                       'JSON: {"0": "...", "1": "..."}')
        for chunk in _chunks(flat, 16):
            payload = {str(i): entry[1][entry[2]].get("item_name")
                       for i, entry in enumerate(chunk)}
            result = _chat_json(system_adds, json.dumps(payload, ensure_ascii=False))
            for i, (r, adds, j) in enumerate(chunk):
                tr = result.get(str(i))
                if isinstance(tr, str) and tr.strip():
                    adds[j]["item_name_traduit"] = tr.strip()
        for r, adds in rows_with_adds:
            r.articles_additionnels = json.dumps(adds, ensure_ascii=False)
            if (r.item_code or r.item_name) not in updated:
                updated.append(r.item_code or r.item_name)

    doc.save(ignore_permissions=True)
    frappe.db.commit()
    return {"updated": updated, "langue": doc.langue_cible}


@frappe.whitelist()
def get_group_paths():
    """{groupe: "Parent > Sous-groupe > Groupe"} — hiérarchie complète des
    groupes d'articles (racine exclue), pour organiser les listes."""
    _guard()
    groups = frappe.get_all("Item Group", fields=["name", "parent_item_group"],
                            order_by="lft")
    parent = {g.name: g.parent_item_group for g in groups}
    roots = {g.name for g in groups if not g.parent_item_group}

    def path(n):
        parts = [n]
        cur = parent.get(n)
        while cur and cur in parent:
            parts.append(cur)
            cur = parent.get(cur)
        return " > ".join(p for p in reversed(parts) if p not in roots) or n

    return {g.name: path(g.name) for g in groups}


# ------------------------------------------- variantes / apparentés / bundles

ITEM_FIELDS = ["name", "item_name", "item_group", "image", "stock_uom",
               "custom_volume_m3", "description", "variant_of", "has_variants",
               "disabled", "brand"]


def _attributes_labels(codes):
    """{item_code: "val1 · val2 · val3"} — valeurs d'attributs de variantes,
    dans l'ordre idx, pour libeller compactement chaque variante."""
    if not codes:
        return {}
    rows = frappe.get_all(
        "Item Variant Attribute",
        filters={"parent": ["in", list(codes)], "parenttype": "Item"},
        fields=["parent", "attribute_value"],
        order_by="parent, idx")
    out = {}
    for r in rows:
        if r.attribute_value:  # les templates ont des lignes sans valeur
            out.setdefault(r.parent, []).append(r.attribute_value)
    return {k: " · ".join(v) for k, v in out.items()}


def _bin_stock(codes):
    """{item_code: stock total (somme Bin.actual_qty)} — même logique que
    prevision_import._stock_map, sans filtre entrepôt."""
    if not codes:
        return {}
    rows = frappe.get_all(
        "Bin", filters={"item_code": ["in", list(codes)]},
        fields=["item_code", "sum(actual_qty) as qty"],
        group_by="item_code")
    return {r.item_code: flt(r.qty) for r in rows}


def _item_brief(code, item_row, attrs, stock):
    return {
        "item_code": code,
        "item_name": item_row.get("item_name") or code,
        "item_group": item_row.get("item_group") or "",
        "image": item_row.get("image") or "",
        "uom": item_row.get("stock_uom") or "",
        "volume_unitaire_m3": flt(item_row.get("custom_volume_m3")),
        "description": item_row.get("description") or "",
        "attributes": attrs.get(code, ""),
        "stock": stock.get(code, 0),
        "is_template": cint(item_row.get("has_variants")),
        "brand": item_row.get("brand") or "",
    }


def _bundles_using(item_code):
    """Product Bundles actifs consommant `item_code`, avec leurs composants."""
    parents = frappe.get_all(
        "Product Bundle Item", filters={"item_code": item_code},
        pluck="parent", distinct=True)
    if not parents:
        return []
    bundles = frappe.get_all(
        "Product Bundle", filters={"name": ["in", parents], "disabled": 0},
        fields=["name", "new_item_code"])
    return _bundle_components({b.name: b for b in bundles})


def _bundle_components(bundle_map):
    """bundle_map = {bundle_name: row(name, new_item_code)} → liste de bundles
    enrichis (attributs du produit fini + composants nommés)."""
    if not bundle_map:
        return []
    comp_rows = frappe.get_all(
        "Product Bundle Item",
        filters={"parent": ["in", list(bundle_map)], "parenttype": "Product Bundle"},
        fields=["parent", "item_code", "qty"], order_by="parent, idx")
    comp_codes = {r.item_code for r in comp_rows}
    finished_codes = {b.new_item_code for b in bundle_map.values() if b.new_item_code}
    names = {}
    for it in frappe.get_all("Item",
                             filters={"name": ["in", list(comp_codes | finished_codes)]},
                             fields=["name", "item_name"]):
        names[it.name] = it.item_name
    attrs = _attributes_labels(finished_codes)
    out = []
    for bname, b in bundle_map.items():
        comps = [{"item_code": r.item_code,
                  "item_name": names.get(r.item_code, r.item_code),
                  "qty": flt(r.qty)}
                 for r in comp_rows if r.parent == bname]
        out.append({
            "bundle": bname,
            "item_code": b.new_item_code or bname,
            "item_name": names.get(b.new_item_code, b.new_item_code or bname),
            "attributes": attrs.get(b.new_item_code, ""),
            "components": comps,
        })
    return out


def _pack_composition(item_code):
    """Si `item_code` est lui-même un produit fini (Product Bundle actif),
    retourne ses composants enrichis (détails article + stock + qty par pack),
    sinon None."""
    pb = frappe.get_all(
        "Product Bundle",
        filters={"disabled": 0, "new_item_code": item_code},
        fields=["name", "new_item_code"], limit=1)
    if not pb:
        return None
    bundles = _bundle_components({pb[0].name: pb[0]})
    comps = bundles[0]["components"] if bundles else []
    codes = [c["item_code"] for c in comps]
    details = {r.name: r for r in frappe.get_all(
        "Item", filters={"name": ["in", codes]}, fields=ITEM_FIELDS)}
    stock = _bin_stock(codes)
    attrs = _attributes_labels(codes)
    out = []
    for c in comps:
        brief = _item_brief(c["item_code"], details.get(c["item_code"], frappe._dict()),
                            attrs, stock)
        brief["item_name"] = brief["item_name"] or c["item_name"]
        brief["qty_per_pack"] = flt(c["qty"])
        out.append(brief)
    return out


@frappe.whitelist()
def get_related_items(item_code):
    """Articles apparentés à `item_code`, pour le picker du formulaire LCI :
    - variants : famille de variantes ERPNext (template inclus) ;
    - siblings : frères par suffixe de code (Acc019-A / Acc019-B…), hors famille ;
    - bundles : produits finis (Product Bundle) qui consomment cet article ;
    - composition : composants du pack si l'article est lui-même un produit fini."""
    _guard()
    it = frappe.db.get_value("Item", item_code, ITEM_FIELDS, as_dict=True)
    if not it:
        frappe.throw(_("Article {0} introuvable.").format(item_code))

    # --- famille de variantes
    variant_codes = []
    root = it.variant_of or (item_code if it.has_variants else None)
    if root:
        variant_codes = [root] + frappe.get_all(
            "Item", filters={"variant_of": root, "disabled": 0}, pluck="name",
            order_by="name")
        variant_codes = [c for c in variant_codes if c != item_code]

    # --- frères par suffixe de code (un seul segment après la racine)
    sibling_codes = []
    if "-" in item_code:
        prefix = item_code.rsplit("-", 1)[0]
        pattern = f"^{re.escape(prefix)}-[^-]+$"
        rows = frappe.db.sql(
            """select name from `tabItem`
               where disabled = 0 and (name = %(prefix)s or name regexp %(re)s)
               order by name""",
            {"prefix": prefix, "re": pattern}, as_dict=True)
        in_family = set(variant_codes) | {item_code, root}
        sibling_codes = [r.name for r in rows if r.name not in in_family]

    all_codes = set(variant_codes) | set(sibling_codes) | {item_code}
    items = {r.name: r for r in frappe.get_all(
        "Item", filters={"name": ["in", list(all_codes)]}, fields=ITEM_FIELDS)}
    attrs = _attributes_labels(all_codes)
    stock = _bin_stock(all_codes)

    def briefs(codes):
        return [_item_brief(c, items[c], attrs, stock) for c in codes if c in items]

    return {
        "item": _item_brief(item_code, items.get(item_code, it), attrs, stock),
        "variants": briefs(variant_codes),
        "siblings": briefs(sibling_codes),
        "bundles": _bundles_using(item_code),
        "composition": _pack_composition(item_code),
    }


@frappe.whitelist()
def estimate_finished_products(docname, include_stock=1, bundle_codes=None,
                               limit=80):
    """Estimation des produits finis assemblables depuis les quantités de la
    liste (+ stock actuel si include_stock). Par bundle candidat :
    assemblables = min sur composants de floor(disponible / qté par unité).
    bundle_codes = JSON list de produits finis ajoutés manuellement."""
    _guard()
    doc = frappe.get_doc(DOCTYPE, docname)
    include_stock = cint(include_stock)

    ordered = {}
    additionnels = {}  # composants embarqués « ADDITIONNEL » : qty/pack × packs
    for r in doc.articles:
        if not r.item_code:
            continue
        ordered[r.item_code] = ordered.get(r.item_code, 0) + flt(r.qty)
        try:
            adds = json.loads(r.articles_additionnels or "[]") or []
        except Exception:
            adds = []
        for a in adds:
            code = a.get("item_code")
            q = flt(a.get("qty_par_pack")) * flt(r.qty)
            if code and q:
                additionnels[code] = additionnels.get(code, 0) + q

    # décomposition : une ligne qui est elle-même un pack (Product Bundle actif)
    # compte comme ses composants × qty — ainsi « 150 packs + 150 membranes
    # additionnelles » donne bien membrane dispo = 300.
    direct = {}
    if ordered:
        for b in frappe.get_all("Product Bundle",
                                filters={"disabled": 0,
                                         "new_item_code": ["in", list(ordered)]},
                                fields=["name", "new_item_code"]):
            direct[b.new_item_code] = {"bundle": b.name,
                                       "qty": ordered.pop(b.new_item_code)}
        if direct:
            comp_rows = frappe.get_all(
                "Product Bundle Item",
                filters={"parent": ["in", [d["bundle"] for d in direct.values()]],
                         "parenttype": "Product Bundle"},
                fields=["parent", "item_code", "qty"])
            by_bundle = {d["bundle"]: d["qty"] for d in direct.values()}
            for c in comp_rows:
                ordered[c.item_code] = ordered.get(c.item_code, 0) \
                    + flt(c.qty) * by_bundle[c.parent]

    for code, q in additionnels.items():  # après décomposition des packs
        ordered[code] = ordered.get(code, 0) + q

    user_extra = json.loads(bundle_codes) if isinstance(bundle_codes, str) and bundle_codes else (bundle_codes or [])
    extra = list(set(user_extra) | set(direct))  # les packs commandés sont toujours affichés

    # candidats : bundles actifs dont >= 1 composant est commandé + sélection manuelle
    parents = set()
    if ordered:
        parents |= set(frappe.get_all(
            "Product Bundle Item", filters={"item_code": ["in", list(ordered)]},
            pluck="parent", distinct=True))
    bundle_rows = []
    if parents:
        bundle_rows += frappe.get_all(
            "Product Bundle", filters={"name": ["in", list(parents)], "disabled": 0},
            fields=["name", "new_item_code"])
    if extra:
        known = {b.name for b in bundle_rows} | {b.new_item_code for b in bundle_rows}
        more = frappe.get_all(
            "Product Bundle",
            filters={"disabled": 0,
                     "new_item_code": ["in", [c for c in extra if c not in known]]},
            fields=["name", "new_item_code"])
        bundle_rows += [b for b in more if b.name not in {x.name for x in bundle_rows}]
    if not bundle_rows:
        return {"bundles": [], "include_stock": include_stock,
                "note": _("Aucun produit fini trouvé pour les articles de cette liste.")}

    bundles = _bundle_components({b.name: b for b in bundle_rows})
    comp_codes = {c["item_code"] for b in bundles for c in b["components"]}
    stock = _bin_stock(comp_codes) if include_stock else {}

    out = []
    for b in bundles:
        buildable = None
        limiting = []
        comps = []
        for c in b["components"]:
            per = flt(c["qty"])
            avail = ordered.get(c["item_code"], 0) + (stock.get(c["item_code"], 0) if include_stock else 0)
            n = int((avail + 1e-9) // per) if per > 0 else None
            comps.append({**c, "per_unit": per,
                          "ordered": ordered.get(c["item_code"], 0),
                          "stock": stock.get(c["item_code"], 0),
                          "available": avail,
                          "in_list": c["item_code"] in ordered,
                          "buildable": n})
            if n is not None and (buildable is None or n < buildable):
                buildable = n
        buildable = buildable or 0
        limiting = [c["item_code"] for c in comps
                    if c["buildable"] is not None and c["buildable"] == buildable]
        for c in comps:  # surplus une fois les N unités montées (rechange)
            c["leftover"] = flt(c["available"]) - buildable * flt(c["per_unit"])
        out.append({**b, "components": comps, "buildable": buildable,
                    "limiting": limiting,
                    "direct_qty": flt(direct.get(b["item_code"], {}).get("qty", 0)),
                    "manual": b["item_code"] in set(user_extra)})

    # packs commandés et sélections manuelles d'abord, le reste par assemblables
    out.sort(key=lambda b: (0 if (b["manual"] or b["direct_qty"]) else 1,
                            -b["buildable"], b["item_code"]))
    total = len(out)
    limit = cint(limit) or 80
    out = out[:limit]
    return {"bundles": out, "include_stock": include_stock,
            "total": total, "shown": len(out),
            "note": _("Estimations indépendantes par produit — les composants "
                      "partagés ne sont pas répartis entre produits.")}


@frappe.whitelist()
def preview_row_export(docname, row_name):
    """Aperçu EXACT de ce que l'export (Excel/PDF) rendra pour une ligne :
    désignation, description et lignes ADDITIONNEL finales (traduction,
    marque, UDM localisée), construit avec les mêmes helpers que l'export.
    Basé sur la dernière sauvegarde du document."""
    _guard()
    doc = frappe.get_doc(DOCTYPE, docname)
    row = next((r for r in doc.articles if r.name == row_name), None)
    if not row:
        frappe.throw(_("Ligne introuvable."))
    L = _labels(doc)
    adds = _adds_map(doc).get(row.name, [])
    lines = []
    for a in adds:
        tot = flt(a.get("qty_par_pack")) * flt(row.qty)
        lines.append(f"+ {L['additional']} : {_add_label(a)} "
                     f"— {flt(a.get('qty_par_pack')):g}/{L['pack']} × {flt(row.qty):g} "
                     f"= {tot:g} {_uom_out(a.get('uom'), L)}")
    return {
        "name": row.item_name_traduit or row.item_name or "",
        "desc": row.description_traduite or _strip_html(row.description),
        "adds": lines,
        "qty": f"{flt(row.qty):g} {_uom_out(row.uom, L)}",
        "lang": doc.langue_cible or "English",
    }


# ---------------------------------------------------------------- exports

def _img_path(file_url):
    """Chemin disque d'un fichier /files/... ou /private/files/..., sinon None.

    Les URLs sont souvent encodées (ex. %22 pour « " » dans les articles en
    pouces : C-10"-PP) alors que le fichier sur disque porte le caractère
    littéral — on essaie donc les deux variantes."""
    if not file_url:
        return None
    from urllib.parse import unquote
    try:
        if file_url.startswith("/private/files/"):
            base = ("private", "files")
        elif file_url.startswith("/files/"):
            base = ("public", "files")
        else:
            return None
        raw = os.path.basename(file_url)
        for candidate in (unquote(raw), raw):
            path = frappe.get_site_path(*base, candidate)
            if os.path.exists(path):
                return path
        return None
    except Exception:
        return None


def _img_thumb(file_url, max_px=220):
    """Vignette JPEG compressée (BytesIO) — allège fortement PDF et Excel."""
    import io as _io

    path = _img_path(file_url)
    if not path:
        return None
    try:
        from PIL import Image as PILImage
        img = PILImage.open(path)
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        img.thumbnail((max_px, max_px))
        buf = _io.BytesIO()
        img.save(buf, format="JPEG", quality=82)
        buf.seek(0)
        return buf
    except Exception:
        return None


def _img_data_uri(file_url):
    """Vignette en data URI base64 (fiable pour wkhtmltopdf)."""
    buf = _img_thumb(file_url)
    if not buf:
        return None
    return "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode()


def _esc(v):
    return frappe.utils.escape_html(v or "")


def _row_adds(row):
    """Articles additionnels embarqués sur une ligne (champ JSON de la ligne)."""
    try:
        return json.loads(row.articles_additionnels or "[]") or []
    except Exception:
        return []


def _add_label(a):
    """Libellé d'un additionnel : désignation (traduite si disponible)
    + marque si présente."""
    name = (a.get("item_name_traduit") or a.get("item_name")
            or a.get("item_code") or "")
    brand = (a.get("brand") or "").strip()
    return f"{name} ({brand})" if brand else name


def _adds_map(doc):
    """{row.name: additionnels} avec la marque complétée depuis la fiche
    Article quand elle manque (additionnels enregistrés avant l'ajout du
    champ brand)."""
    rows_adds = {r.name: _row_adds(r) for r in doc.articles}
    missing = {a["item_code"] for adds in rows_adds.values() for a in adds
               if a.get("item_code") and not a.get("brand")}
    if missing:
        brands = {i.name: (i.brand or "") for i in frappe.get_all(
            "Item", filters={"name": ["in", list(missing)]},
            fields=["name", "brand"])}
        for adds in rows_adds.values():
            for a in adds:
                if not a.get("brand"):
                    a["brand"] = brands.get(a.get("item_code"), "")
    return rows_adds


# libellés d'export par langue cible — le document entier (titres, en-têtes,
# mentions ADDITIONNEL) sort dans la langue de la cotation.
EXPORT_LABELS = {
    "Français": {
        "title": "Demande de cotation", "ref": "Référence", "date": "Date",
        "supplier": "Fournisseur", "language": "Langue",
        "image": "Image", "code": "Code article", "designation": "Désignation",
        "description": "Description", "qty": "Quantité", "uom": "UDM",
        "unit_vol": "Volume unitaire (m³)", "line_vol": "Volume ligne (m³)",
        "unit_price": "Prix unitaire", "line_total": "Total",
        "total": "TOTAL", "articles": "articles",
        "total_volume": "Volume total estimé",
        "additional": "ADDITIONNEL", "pack": "pack", "pcs": "Pièce",
    },
    "English": {
        "title": "Quotation Request", "ref": "Reference", "date": "Date",
        "supplier": "Supplier", "language": "Language",
        "image": "Image", "code": "Item code", "designation": "Item name",
        "description": "Description", "qty": "Quantity", "uom": "UOM",
        "unit_vol": "Unit volume (m³)", "line_vol": "Line volume (m³)",
        "unit_price": "Unit price", "line_total": "Total",
        "total": "TOTAL", "articles": "items",
        "total_volume": "Estimated total volume",
        "additional": "ADDITIONAL", "pack": "pack", "pcs": "Pcs",
    },
    "Deutsch": {
        "title": "Angebotsanfrage", "ref": "Referenz", "date": "Datum",
        "supplier": "Lieferant", "language": "Sprache",
        "image": "Bild", "code": "Artikelnummer", "designation": "Bezeichnung",
        "description": "Beschreibung", "qty": "Menge", "uom": "Einheit",
        "unit_vol": "Stückvolumen (m³)", "line_vol": "Zeilenvolumen (m³)",
        "unit_price": "Stückpreis", "line_total": "Gesamt",
        "total": "GESAMT", "articles": "Artikel",
        "total_volume": "Geschätztes Gesamtvolumen",
        "additional": "ZUSÄTZLICH", "pack": "Pack", "pcs": "Stk",
    },
    "العربية": {
        "title": "طلب عرض أسعار", "ref": "المرجع", "date": "التاريخ",
        "supplier": "المورّد", "language": "اللغة",
        "image": "الصورة", "code": "رمز الصنف", "designation": "التسمية",
        "description": "الوصف", "qty": "الكمية", "uom": "الوحدة",
        "unit_vol": "حجم الوحدة (م³)", "line_vol": "حجم السطر (م³)",
        "unit_price": "سعر الوحدة", "line_total": "المجموع",
        "total": "المجموع", "articles": "أصناف",
        "total_volume": "الحجم الإجمالي التقديري",
        "additional": "إضافي", "pack": "علبة", "pcs": "قطعة",
    },
    "Español": {
        "title": "Solicitud de cotización", "ref": "Referencia", "date": "Fecha",
        "supplier": "Proveedor", "language": "Idioma",
        "image": "Imagen", "code": "Código", "designation": "Denominación",
        "description": "Descripción", "qty": "Cantidad", "uom": "UdM",
        "unit_vol": "Volumen unitario (m³)", "line_vol": "Volumen línea (m³)",
        "unit_price": "Precio unitario", "line_total": "Total",
        "total": "TOTAL", "articles": "artículos",
        "total_volume": "Volumen total estimado",
        "additional": "ADICIONAL", "pack": "pack", "pcs": "Uds",
    },
}


def _uom_out(uom, L):
    """UDM localisée à l'export — traduit le « Pièce » dominant du catalogue."""
    return L["pcs"] if (uom or "").strip().lower() in ("pièce", "piece", "piéce", "pc", "pcs") else (uom or "")


def _labels(doc):
    return EXPORT_LABELS.get(doc.langue_cible or "English", EXPORT_LABELS["English"])


@frappe.whitelist()
def download_pdf(docname):
    from frappe.utils.pdf import get_pdf

    _guard()
    doc = frappe.get_doc(DOCTYPE, docname)
    rtl = doc.langue_cible in RTL_LANGS
    dir_attr = ' dir="rtl"' if rtl else ""
    L = _labels(doc)

    supplier = ""
    if doc.fournisseur:
        supplier = frappe.db.get_value("Supplier", doc.fournisseur, "supplier_name") or doc.fournisseur

    adds_map = _adds_map(doc)
    rows_html = ""
    for i, r in enumerate(doc.articles, 1):
        img = _img_data_uri(r.image)
        img_html = f'<img src="{img}" style="max-width:90px;max-height:90px;">' if img else ""
        name = r.item_name_traduit or r.item_name or ""
        desc = r.description_traduite or _strip_html(r.description)

        # additionnels embarqués : en ROUGE, quantité = multiplication par pack
        adds = adds_map.get(r.name, [])
        adds_desc = ""
        adds_qty = ""
        adds_img = ""
        for a in adds:
            tot = flt(a.get("qty_par_pack")) * flt(r.qty)
            aname = _esc(_add_label(a))
            auom = _esc(_uom_out(a.get("uom"), L))
            aimg = _img_data_uri(a.get("image"))
            if aimg:
                adds_img += (f'<img src="{aimg}" style="max-width:34px;max-height:34px;'
                             f'border:1px solid #d99;border-radius:3px;margin:2px;">')
            adds_desc += (f'<div style="color:#b00020;background:#fdf0f0;'
                          f'border:1px solid #e8b0b0;border-radius:4px;'
                          f'padding:3px 6px;margin-top:4px;">'
                          f'<b>+ {L["additional"]}</b> : {aname} — '
                          f'{flt(a.get("qty_par_pack")):g}/{L["pack"]} × {flt(r.qty):g} '
                          f'= <b>{tot:g} {auom}</b></div>')
            adds_qty += (f'<div style="color:#b00020;font-weight:bold;">'
                         f'+{tot:g} {auom}</div>')
        rows_html += f"""
        <tr>
          <td style="text-align:center;">{i}</td>
          <td style="text-align:center;">{img_html}{('<br>' + adds_img) if adds_img else ''}</td>
          <td><span{dir_attr}><b>{_esc(name)}</b></span></td>
          <td{dir_attr}>{_esc(desc).replace(chr(10), '<br>')}{adds_desc}</td>
          <td style="text-align:right;">{flt(r.qty):g} {_esc(_uom_out(r.uom, L))}{adds_qty}</td>
          <td style="text-align:right;">{flt(r.volume_ligne_m3):.3f}</td>
        </tr>"""

    html = f"""
    <html><head><meta charset="utf-8"><style>
      body {{ font-family: Helvetica, Arial, sans-serif; font-size: 11px; color: #222; }}
      h1 {{ font-size: 18px; margin-bottom: 2px; }}
      .meta {{ color: #555; margin-bottom: 14px; }}
      table {{ width: 100%; border-collapse: collapse; }}
      th {{ background: #f0f2f5; text-align: left; padding: 6px 8px; font-size: 10px;
            text-transform: uppercase; border: 1px solid #d5dae1; }}
      td {{ padding: 6px 8px; border: 1px solid #d5dae1; vertical-align: top; }}
      .tot {{ margin-top: 12px; font-size: 12px; font-weight: bold; text-align: right; }}
    </style></head><body>
      <h1>{L["title"]} — {_esc(doc.titre)}</h1>
      <div class="meta">
        {L["ref"]} : {_esc(doc.name)} &nbsp;·&nbsp; {L["date"]} : {_esc(str(doc.date_commande or nowdate()))}
        {"&nbsp;·&nbsp; " + L["supplier"] + " : <b>" + _esc(supplier) + "</b>" if supplier else ""}
        &nbsp;·&nbsp; {L["language"]} : {_esc(doc.langue_cible or "")}
      </div>
      <table>
        <thead><tr>
          <th style="width:24px;">#</th><th style="width:100px;">{L["image"]}</th>
          <th style="width:150px;">{L["designation"]}</th><th>{L["description"]}</th>
          <th style="width:80px;text-align:right;">{L["qty"]}</th>
          <th style="width:70px;text-align:right;">{L["line_vol"]}</th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
      <div class="tot">
        {cint(doc.nb_articles)} {L["articles"]} &nbsp;·&nbsp; {L["total_volume"]} :
        {flt(doc.volume_total_m3):.3f} m³
      </div>
    </body></html>"""

    pdf = get_pdf(html)
    frappe.local.response.filename = f"{doc.name}-cotation.pdf"
    frappe.local.response.filecontent = pdf
    frappe.local.response.type = "download"


@frappe.whitelist()
def download_excel(docname):
    """Excel de cotation : images embarquées, désignation/description TRADUITES
    (repli sur l'original), colonnes prix à remplir par le fournisseur."""
    import io

    from openpyxl import Workbook
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.styles.colors import Color
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.units import pixels_to_EMU

    _guard()
    doc = frappe.get_doc(DOCTYPE, docname)
    L = _labels(doc)
    DESC_PT = 8  # petite police pour la description
    small = InlineFont(sz=DESC_PT)
    red = InlineFont(b=True, color=Color(rgb="FFB00020"), sz=DESC_PT)

    wb = Workbook()
    ws = wb.active
    ws.title = L["title"][:31]
    if doc.langue_cible in RTL_LANGS:
        ws.sheet_view.rightToLeft = True

    header = ["#", L["image"], L["designation"], L["description"],
              L["qty"], L["uom"], L["unit_vol"], L["line_vol"],
              L["unit_price"], L["line_total"]]
    widths = [4, 14, 36, 58, 10, 8, 12, 12, 12, 12]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.append([f"{L['title']} — {doc.titre} ({doc.name})"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"{L['date']} : {doc.date_commande or nowdate()}"
               + (f" — {L['supplier']} : {doc.fournisseur}" if doc.fournisseur else "")
               + f" — {L['language']} : {doc.langue_cible or ''}"])
    ws.append([])
    ws.append(header)
    hrow = ws.max_row
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=hrow, column=c)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill("solid", fgColor="F0F2F5")

    def _embed(file_url, row_no, max_side, col_px_off, row_px_off):
        """Insère une image dans la colonne B avec un décalage en pixels."""
        buf_img = _img_thumb(file_url, max_px=160)
        if not buf_img:
            return False
        try:
            img = XLImage(buf_img)
            ratio = min(max_side / (img.width or max_side),
                        max_side / (img.height or max_side), 1.0)
            w = int((img.width or max_side) * ratio)
            h = int((img.height or max_side) * ratio)
            marker = AnchorMarker(col=1, colOff=pixels_to_EMU(col_px_off),
                                  row=row_no - 1, rowOff=pixels_to_EMU(row_px_off))
            img.anchor = OneCellAnchor(
                _from=marker,
                ext=XDRPositiveSize2D(pixels_to_EMU(w), pixels_to_EMU(h)))
            ws.add_image(img)
            return True
        except Exception:
            return False

    wrap = Alignment(wrap_text=True, vertical="center")
    vcenter = Alignment(vertical="center")
    adds_map = _adds_map(doc)
    for i, r in enumerate(doc.articles, 1):
        name = r.item_name_traduit or r.item_name or ""
        desc = r.description_traduite or _strip_html(r.description)
        adds = adds_map.get(r.name, [])

        ws.append([i, "", name, "", flt(r.qty),
                   _uom_out(r.uom, L),
                   flt(r.volume_unitaire_m3), flt(r.volume_ligne_m3), "", ""])
        row_no = ws.max_row

        # description concaténée avec les ADDITIONNELS (texte riche, en rouge),
        # le tout en petite police (DESC_PT)
        dcell = ws.cell(row=row_no, column=4)
        parts = [TextBlock(small, desc)] if desc else []
        for a in adds:
            tot = flt(a.get("qty_par_pack")) * flt(r.qty)
            parts.append(TextBlock(
                red,
                f"\n+ {L['additional']} : {_add_label(a)} "
                f"— {flt(a.get('qty_par_pack')):g}/{L['pack']} × {flt(r.qty):g} "
                f"= {tot:g} {_uom_out(a.get('uom'), L)}"))
        dcell.value = CellRichText(*parts) if parts else ""
        for c in range(1, len(header) + 1):  # toute la ligne centrée verticalement
            ws.cell(row=row_no, column=c).alignment = wrap if c in (3, 4) else vcenter

        # hauteur : description complète visible (≈10 pt par ligne en police 8)
        nb_lines = (desc.count("\n") + 1 if desc else 1) + len(adds)
        est_lines = max(nb_lines, int(len(desc) / 75) + 1 if desc else 1)
        height = max(52, est_lines * 10 + 6)
        if adds:
            height = max(height, 52 + 30)  # place pour les mini-images
        ws.row_dimensions[row_no].height = height

        # image principale + mini-images des additionnels en dessous
        _embed(r.image, row_no, 64, 2, 2)
        col_off = 2
        for a in adds:
            if a.get("image") and _embed(a["image"], row_no, 32, col_off, 70):
                col_off += 36

    ws.append([])
    ws.append(["", "", L["total"], "", "", "", "", flt(doc.volume_total_m3), "", ""])
    ws.cell(row=ws.max_row, column=3).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=8).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    frappe.response["filename"] = f"{doc.name}-cotation.xlsx"
    frappe.response["filecontent"] = buf.getvalue()
    frappe.response["type"] = "binary"
